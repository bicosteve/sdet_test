import openpyxl


class ExcelUtility:
    @staticmethod
    def get_row_count(file, sheet_name):
        workbook = openpyxl.load_workbook(file)
        sheet = workbook[sheet_name]
        return sheet.max_row

    @staticmethod
    def read_data(file, sheet_name, row_number, column_number):
        workbook = openpyxl.load_workbook(file)
        sheet = workbook[sheet_name]
        return sheet.cell(row=row_number, column=column_number).value
